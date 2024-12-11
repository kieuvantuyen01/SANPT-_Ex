import sys
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from zipfile import BadZipFile
from typing import List

# from pysat.formula import CNF
from pysat.solvers import Glucose3, Solver
from itertools import product
import time
from threading import Timer, Thread, Event
import os
import ast
from pypblib import pblib
from pypblib.pblib import PBConfig, Pb2cnf

sat_solver = Glucose3
time_budget = 600  # Set your desired time budget in seconds
type = "es3_improved_pb_block"
id_counter = 1
id_variable: int

# Open the log file in append mode
log_file = open('console.log', 'a')

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'out/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

    print_to_console_and_log(f"Result added to Excel file: {os.path.abspath(excel_file_path)}\n")

# Define a custom print function that writes to both console and log file
def print_to_console_and_log(*args, **kwargs):
    print(*args, **kwargs)
    print(*args, file = log_file, **kwargs)
    log_file.flush()

def exactly_k(var: List[int], k):
    global id_variable

    pbConfig = PBConfig()
    pbConfig.set_PB_Encoder(pblib.PB_BDD)

    # Create a Pb2cnf object
    pb2 = Pb2cnf(pbConfig)

    # Create a list to hold the formula
    formula = []

    # Create a list to hold the weights are 1
    weights = [1] * len(var)

    # Encode the AtLeastK and AtMostK constraints
    # max_var = pb2.encode_at_least_k(var, k, formula, id_variable + 1)
    # max_var = pb2.encode_at_most_k(var, k, formula, max_var + 1)

    # encode_both()
    max_var = pb2.encode_both(weights, var, k, k, formula, id_variable + 1)

    for clause in formula:
        sat_solver.add_clause(clause)
        # print(f"Added clause: {clause}")

    # Update the global variable id_variable based on the new variables introduced by the encoding
    id_variable = max_var

def check_overlap(task1, task2):
    # Suppose: task1 = (r1, e1, d1), task2 = (r2, e2, d2)
    # r1_min = r1, r1_max = d1 - e1, d1_min = r1 + e1, d1_max = d1
    # r2_min = r2, r2_max = d2 - e2, d2_min = r2 + e2, d2_max = d2
    # task1 and task2 are overlapped if: 
    # 1. d2_min >= r1_max and r2_max <= d1_min
    # 2. d1_min >= r2_max and r1_max <= d2_min
    # => r2 + e2 >= d1 - e1 and d2 - e2 <= r1 + e1 or r1 + e1 >= d2 - e2 and d1 - e1 <= r2 + e2
    if task2[0] + task2[1] > task1[2] - task1[1] and task2[2] - task2[1] < task1[0] + task1[1]:
        return True
    if task1[0] + task1[1] > task2[2] - task2[1] and task1[2] - task1[1] < task2[0] + task2[1]:
        return True
    return False

def block_encoding(X, k, var_index):
    n = len(X) - 1
    clauses = []
    aux_vars = {}  # Keep track of auxiliary variables
    current_var_index = var_index

    ra_clauses, current_var_index, ra_final = encode_all_zero_block(X, n, k, current_var_index)
    clauses.extend(ra_clauses)

    rb_clauses, current_var_index, rb_final = encode_left_all_one_block(X, n, k, current_var_index)
    clauses.extend(rb_clauses)

    rc_clauses, current_var_index, rc_final = encode_right_all_one_block(X, n, k, current_var_index)
    clauses.extend(rc_clauses)

    # Add the main implication clause
    # For each implication in the formula
    # First line
    # X1 -> Ra,1 ^ Rb,1 ^ Rc,1
    if ra_final and len(ra_final) > 0:
        clauses.append([-X[1], ra_final[0]])
    if rb_final and len(rb_final) > 0: 
        if  rb_final and rb_final[0]:
            clauses.append([-X[1], rb_final[0]])
    if rc_final and len(rc_final) > 0: 
        clauses.append([-X[1], rc_final[0]])
    for i in range(2, n-k+1):
        # -X[i-1] ^ X[i] -> Ra,i ^ Rb,i ^ Rc,i
        clauses.append([X[i-1], -X[i], ra_final[i-1]])
        if rb_final and len(rb_final) >= i and rb_final[i-1]:
            clauses.append([X[i-1], -X[i], rb_final[i-1]])
        if rc_final and len(rc_final) >= i:
            clauses.append([X[i-1], -X[i], rc_final[i-1]])

    # Last line
    if n > k and rb_final and len(rb_final) >= n-k+1:
        if rb_final[n-k]:
            clauses.append([X[n-k], -X[n-k+1], rb_final[n-k]])
    if n > k and rc_final and len(rc_final) >= n-k+1: 
        clauses.append([X[n-k], -X[n-k+1], rc_final[n-k]])
    
    return clauses, current_var_index

def encode_all_zero_block(X, n, k, var_index):
    """Encode block All Zero using auxiliary variables"""
    clauses = []
    r_vars = []  # Store all r variables

    # Add the last variable
    if n >= k+1:
        r_vars.append(-X[n])
    
    # First clause: -Xn-1 ^ -Xn -> R1
    if n >= k+2:
        r1 = var_index
        clauses.append([X[-2], X[-1], r1])
        clauses.append([-X[-2], -r1])
        clauses.append([-X[-1], -r1])
        r_vars.append(r1)
        current_r = r1  # Initialize current_r
    
    # For remaining variables
    for i in range(n-2, k, -1):
        new_r = var_index + 1
        var_index += 1
        # -Xn-m-1 ^ Rm -> Rm+1
        clauses.append([X[i], -current_r, new_r])
        # Xn-m-1 -> -Rm+1
        clauses.append([-X[i], -new_r])
        # -Rm -> -Rm+1
        clauses.append([current_r, -new_r])
        current_r = new_r
        r_vars.append(new_r)
    
    r_vars.reverse()
    return clauses, var_index + 1, r_vars

# def encode_all_zero_block(X, n, k, var_index):
#     """Encode block All Zero using auxiliary variables"""
#     clauses = []
#     r_vars = []  # Store all r variables

#     # Add the last variable
#     if n >= k+1:
#         r_vars.append(-X[n])
    
#     # First clause: -Xn-1 ^ -Xn -> R1
#     if n >= k+2:
#         r1 = var_index
#         clauses.append([X[-2], X[-1], r1])
#         clauses.append([-X[-2], -r1])
#         clauses.append([-X[-1], -r1])
#         r_vars.append(r1)
#         current_r = r1  # Initialize current_r
    
#     # For remaining variables
#     for i in range(n-2, k, -1):
#         new_r = var_index + 1
#         var_index += 1
#         # -Xn-m-1 ^ Rm -> Rm+1
#         clauses.append([X[i], -current_r, new_r])
#         # Xn-m-1 -> -Rm+1
#         clauses.append([-X[i], -new_r])
#         # -Rm -> -Rm+1
#         clauses.append([current_r, -new_r])
#         current_r = new_r
#         r_vars.append(new_r)
    
#     r_vars.reverse()
#     return clauses, var_index + 1, r_vars

def encode_left_all_one_block(X, n, k, var_index):
    clauses = []
    r_vars = []
    start_id = 2
    end_id = start_id + k - 3
    
    while (start_id < n and end_id < n):
        r_vars_tmp = []
        # Add the last variable
        r_vars_tmp.append(X[end_id])

        if start_id > end_id:
            break
        if start_id == end_id:
            start_id = start_id + k - 1
            end_id = start_id + k - 3
            r_vars.extend(r_vars_tmp)
            if (start_id < n and end_id < n):
                r_vars.append(0)
            continue
            
        r1 = var_index
        clauses.append([-X[end_id-1], -X[end_id], r1])
        clauses.append([X[end_id-1], -r1])
        clauses.append([X[end_id], -r1])
        r_vars_tmp.append(r1)
        current_r = r1  # Update current_r

        for i in range(end_id-2, start_id-1, -1):
            new_r = var_index + 1
            var_index += 1
            clauses.append([-X[i], -current_r, new_r])
            clauses.append([X[i], -new_r])
            clauses.append([current_r, -new_r])
            current_r = new_r
            r_vars_tmp.append(new_r)

        start_id = start_id + k - 1
        end_id = start_id + k - 3
        r_vars_tmp.reverse()  # Reverse in place
        r_vars.extend(r_vars_tmp)  # Then extend
        if (start_id < n and end_id < n):
            var_index += 1
            r_vars.append(0)
    
    return clauses, var_index + 1, r_vars
    
def encode_right_all_one_block(X, n, k, var_index):
    clauses = []
    r_vars = []
    start_id = k
    end_id = start_id + k - 2

    while (start_id <= n and end_id <= n):
        # Add the first variable
        if k > 1:
            r_vars.append(X[start_id])

        if start_id > end_id:
            break
        if start_id == end_id:
            # r_vars.append(X[end_id])
            start_id = start_id + k - 1
            end_id = start_id + k - 2
            # if (start_id <= n):
            #     var_index += 1
            continue

        r1 = var_index
        clauses.append([-X[start_id], -X[start_id+1], r1])
        clauses.append([X[start_id], -r1])
        clauses.append([X[start_id+1], -r1])
        r_vars.append(r1)
        current_r = r1  # Update current_r

        for i in range(start_id+2, end_id+1):
            new_r = var_index + 1
            var_index += 1
            clauses.append([-X[i], -current_r, new_r])
            clauses.append([X[i], -new_r])
            clauses.append([current_r, -new_r])
            current_r = new_r
            r_vars.append(new_r)

        start_id = start_id + k - 1
        end_id = start_id + k - 2
        if (start_id <= n):
            var_index += 1

    if start_id <= n and end_id > n:
        r_vars.append(X[start_id])
        if start_id+1 <= n:
            r1 = var_index
            clauses.append([-X[start_id], -X[start_id+1], r1])
            clauses.append([X[start_id], -r1])
            clauses.append([X[start_id+1], -r1])
            r_vars.append(r1)
            current_r = r1  # Update current_r

        for i in range(start_id+2, n+1):
            new_r = var_index + 1
            var_index += 1
            clauses.append([-X[i], -current_r, new_r])
            clauses.append([X[i], -new_r])
            clauses.append([current_r, -new_r])
            current_r = new_r
            r_vars.append(new_r)

    return clauses, var_index + 1, r_vars

def encode_problem_es3(tasks, resources):
    global id_variable
    max_time = max(task[2] for task in tasks)

    # Variables u[i][j] for task i accessing resource j
    u = [[i * resources + j + 1 for j in range(resources)] for i in range(len(tasks))]

    # Variables z[i][t] for task i accessing some resource at time t
    z = [[len(tasks) * resources + i * max_time + t + 1 for t in range(tasks[i][2])] for i in range(len(tasks))]

    # Calculate id_variable
    id_variable = len(tasks) * resources + len(tasks) * max_time

    # Overlapping: check each pair of tasks to see if they are overlap time, u_i1j -> -u_i2j
    for i in range(len(tasks)):
        for ip in range(i + 1, len(tasks)):
            if check_overlap(tasks[i], tasks[ip]):
                for j in range(resources):
                    sat_solver.add_clause([-u[i][j], -u[ip][j]])
                    # print(f"Added clause D0: -u{i+1}{j+1} -u{ip+1}{j+1}")

    # Symmetry breaking 1: Assign the tasks to resources if have r_max <= d_min (min of all tasks)
    d_min = min(task[2] for task in tasks)
    fixed_tasks = []
    for i in range(len(tasks)):
        if tasks[i][2] - tasks[i][1] <= d_min:
            fixed_tasks.append(i)
    # Assign each task in fixed_tasks to a resource
    for j, i in enumerate(fixed_tasks):
        if j < resources:
            sat_solver.add_clause([u[i][j]])
        # print(f"Added clause S1: u{i+1}{j+1}")
    
    # Symmetry breaking 2: if each task i has t in range(r_max, d_min), then z[i][t] = True
    # for j in range(resources):
    for i in range(len(tasks)):
        for t in range(tasks[i][2] - tasks[i][1], tasks[i][0] + tasks[i][1]):
            sat_solver.add_clause([z[i][t]])
            # print(f"Added clause S2: -u{i+1}{j+1}, z{i+1}{t}")

    # # D1: Task i should not access two resources at the same time
    # for i in range(len(tasks)):
    #     for j in range(resources):
    #         for jp in range(j + 1, resources):
    #             sat_solver.add_clause([-u[i][j], -u[i][jp]])
    #             # print(f"Added clause D1: -u{i+1}{j+1} -u{i+1}{jp+1}")

    # # D2: Each task must get some resource
    # for i in range(len(tasks)):
    #     # sat_solver.add_clause([u[i][j] for j in range(resources)])
    #     # print(f"Added clause: u{i}0 u{i}1")
    #     clause = []
    #     clause_str = []
    #     for j in range(resources):
    #         clause.append(u[i][j])
    #         clause_str.append(f"u{i+1}{j+1}")
    #     sat_solver.add_clause(clause)
    #     # print(f"Added clause D2: {clause_str}")

    # D1, D2: Each task should access exactly one resource
    for i in range(len(tasks)):
        u_list = []
        for j in range(resources):
            u_list.append(u[i][j])
        # print(f"u_list: {u_list}")
        exactly_k(u_list, 1)

     # D3: A resource can only be held by one task at a time
    for i in range(len(tasks)):
        for ip in range(i + 1, len(tasks)):
            for j in range(resources):
                for t in range(tasks[i][0], min(tasks[i][2], tasks[ip][2])):
                    sat_solver.add_clause([-z[i][t], -u[i][j], -z[ip][t], -u[ip][j]])
                    # print(f"Added clause D3: -z{i+1}{t} -u{i+1}{j+1} -z{ip+1}{t} -u{ip+1}{j+1}")
    
    for i in range(len(tasks)):
        clause = []
        clause_str = []
        for t in range(tasks[i][0], tasks[i][2] - tasks[i][1] + 1):
            clause.append(z[i][t])
            clause_str.append(f"z{i+1}{t}")
        sat_solver.add_clause(clause)
        # print(f"Added clause C3: {clause_str}")

    # check each pair z_i^t and z_i^t+1, if u_ij ^ -z_i^t ^ z_i^t+1 -> ^ z_list[j]
    for i in range(len(tasks)):
        # for t in range(tasks[i][0] + 1, tasks[i][0] + tasks[i][1]):
        #     sat_solver.add_clause([-z[i][tasks[i][0]], z[i][t]])
        #     # print(f"Added clause C41: -z{i+1}{tasks[i][0]} z{i+1}{t}")

        # for t in range (tasks[i][0] + tasks[i][1], tasks[i][2]):
        #     sat_solver.add_clause([-z[i][tasks[i][0]], -z[i][t]])
        #     # print(f"Added clause C42: -z{i+1}{tasks[i][0]} -z{i+1}{t}")

        # for t in range(tasks[i][0], tasks[i][2] - tasks[i][1]):
        #     for tpp in range(t+1, t + tasks[i][1] + 1):
        #         if tpp < max_time:
        #             sat_solver.add_clause([z[i][t], -z[i][t+1], z[i][tpp]])
        #             # print(f"Added clause C51: z{i+1}{t}, -z{i+1}{t+1}, z{i+1}{tpp}")

        #     for tpp in range(t + tasks[i][1] + 1, tasks[i][2]):
        #         if tpp < max_time:
        #             sat_solver.add_clause([z[i][t], -z[i][t+1], -z[i][tpp]])
        #             # print(f"Added clause C52: z{i+1}{t}, -z{i+1}{t+1}, -z{i+1}{tpp}")

        X = [0]
        for t in range(tasks[i][0], tasks[i][2]):
            X.append(z[i][t])
        k = tasks[i][1]
        clauses, final_var_index = block_encoding(X, k, id_variable)
        for clause in clauses:
            sat_solver.add_clause(clause)
            # print(f"Added clause C4: {clause}")
        id_variable = final_var_index

    # sat_solver.add_clause([z[1][3]])
    return u, z

def validate_solution(tasks, model, u, z, resources):
    task_resource = {}
    task_times = {}
    resource_usage = {j: [] for j in range(resources)}

    for i, task in enumerate(tasks):
        for j in range(resources):
            if model[u[i][j] - 1] > 0:
                task_resource[i] = j
        
        task_times[i] = [t for t in range(task[0], task[2]) if model[z[i][t] - 1] > 0]
        
        if task_resource.get(i) is not None:
            resource_usage[task_resource[i]].extend(task_times[i])

    # Check constraints
    for i, task in enumerate(tasks):
        # Check if task is assigned to exactly one resource
        if i not in task_resource:
            print_to_console_and_log(f"Error: Task {i} is not assigned to any resource")
            return False

        # Check if task starts after its release time
        if task_times[i][0] < task[0]:
            print_to_console_and_log(f"Error: Task {i+1} starts before its release time")
            return False

        # Check if task finishes before its deadline
        if task_times[i][-1] >= task[2]:
            print_to_console_and_log(f"Error: Task {i+1} finishes after its deadline")
            return False

        # Check if task execution is continuous and matches the execution time
        if len(task_times[i]) != task[1] or any(task_times[i][j+1] - task_times[i][j] != 1 for j in range(len(task_times[i])-1)):
            print_to_console_and_log(f"Error: Task {i+1} execution is not continuous or doesn't match execution time")
            return False

    # Check if any resource is used by multiple tasks at the same time
    for j, times in resource_usage.items():
        if len(times) != len(set(times)):
            print_to_console_and_log(f"Error: Resource {j+1} is used by multiple tasks at the same time")
            return False

    print_to_console_and_log("Solution is valid!")
    return True

def solve_with_timeout(tasks, resources, result_container, finished_event):
    global sat_solver
    sat_solver = Glucose3()
    
    try:
        u, z = encode_problem_es3(tasks, resources)
        result = sat_solver.solve()
        
        if result:
            model = sat_solver.get_model()
            result_container['status'] = 'SAT'
            result_container['model'] = model
            result_container['u'] = u
            result_container['z'] = z
        else:
            result_container['status'] = 'UNSAT'
            
    except Exception as e:
        result_container['status'] = 'ERROR'
        result_container['error'] = str(e)
    
    finished_event.set()

def solve_es3(tasks, resources):
    global sat_solver
    
    result_container = {}
    finished_event = Event()
    
    start_time = time.time()
    solver_thread = Thread(target=solve_with_timeout, args=(tasks, resources, result_container, finished_event))
    solver_thread.start()
    
    # Wait for either completion or timeout
    finished = finished_event.wait(timeout=time_budget)
    solve_time = time.time() - start_time
    
    if not finished:
        sat_solver.interrupt()
        solver_thread.join()  # Wait for thread to clean up
        sat_solver.delete()
        return "Time out", solve_time, 0, 0
        
    if result_container.get('status') == 'SAT':
        model = result_container['model']
        u = result_container['u']
        z = result_container['z']
        
        print("SAT")
        for i in range(len(tasks)):
            for j in range(resources):
                if model[u[i][j] - 1] > 0:
                    print_to_console_and_log(f"Task {i+1} is assigned to resource {j+1}")
            for t in range(tasks[i][0], tasks[i][2]):
                if model[z[i][t] - 1] > 0:
                    print_to_console_and_log(f"Task {i+1} is accessing a resource at time {t}")
        
        if not validate_solution(tasks, model, u, z, resources):
            sys.exit(1)
        
        number_of_variables = sat_solver.nof_vars()
        number_of_clauses = sat_solver.nof_clauses()
        sat_solver.delete()
        return "SAT", solve_time, number_of_variables, number_of_clauses
    
    elif result_container.get('status') == 'UNSAT':
        print_to_console_and_log("UNSAT")
        number_of_variables = sat_solver.nof_vars()
        number_of_clauses = sat_solver.nof_clauses()
        sat_solver.delete()
        return "UNSAT", solve_time, number_of_variables, number_of_clauses
    
    else:
        print_to_console_and_log(f"Error: {result_container.get('error')}")
        sat_solver.delete()
        return "ERROR", solve_time, 0, 0
    
def process_input_files(input_folder, resources=200):
    global id_counter, type

    # results = {}
    for filename in os.listdir(input_folder):
        if filename.endswith(".txt"):
            file_path = os.path.join(input_folder, filename)
            with open(file_path, 'r') as f:
                num_tasks = int(f.readline().strip())
                tasks = ast.literal_eval(f.readline().strip())
                print(f"tasks: {tasks}")

            print_to_console_and_log(f"Processing {filename}...")
            # res, solve_time, num_variables, num_clauses = solve_es3(tasks, num_tasks)
            res, solve_time, num_variables, num_clauses = solve_es3(tasks, resources)
            result_dict = {
                "ID": id_counter,
                "Problem": os.path.basename(filename),
                "Type": type,
                "Time": solve_time,
                "Result": res,
                "Variables": num_variables,
                "Clauses": num_clauses
            }
            write_to_xlsx(result_dict)
            id_counter += 1

    # return results

# Main execution
input_folder = "input/" + sys.argv[1]
# input_folder = "input/small"
process_input_files(input_folder)

log_file.close()
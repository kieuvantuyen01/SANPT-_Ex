import sys
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from zipfile import BadZipFile
from ortools.linear_solver import pywraplp
from itertools import product
import os
import ast
import time
from threading import Thread, Event

time_budget = 600  # Set your desired time budget in seconds
type = "es3_improved_ortools_mip"
id_counter = 1

# Open the log file in append mode
log_file = open('console.log', 'a')

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'out/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

    print_to_console_and_log(f"Result added to Excel file: {os.path.abspath(excel_file_path)}\n")


# Define a custom print function that writes to both console and log file
def print_to_console_and_log(*args, **kwargs):
    print(*args, **kwargs)
    print(*args, file = log_file, **kwargs)
    log_file.flush()

def check_overlap(task1, task2):
    # Suppose: task1 = (r1, e1, d1), task2 = (r2, e2, d2)
    # r1_min = r1, r1_max = d1 - e1, d1_min = r1 + e1, d1_max = d1
    # r2_min = r2, r2_max = d2 - e2, d2_min = r2 + e2, d2_max = d2
    # task1 and task2 are overlapped if: 
    # 1. d2_min >= r1_max and r2_max <= d1_min
    # 2. d1_min >= r2_max and r1_max <= d2_min
    # => r2 + e2 >= d1 - e1 and d2 - e2 <= r1 + e1 or r1 + e1 >= d2 - e2 and d1 - e1 <= r2 + e2
    if task2[0] + task2[1] > task1[2] - task1[1] and task2[2] - task2[1] < task1[0] + task1[1]:
        return True
    if task1[0] + task1[1] > task2[2] - task2[1] and task1[2] - task1[1] < task2[0] + task2[1]:
        return True
    return False

def encode_problem_es3(tasks, resources):
    solver = pywraplp.Solver.CreateSolver('SCIP')
    if not solver:
        return None

    max_time = max(task[2] for task in tasks)

    # Variables u[i][j] for task i accessing resource j
    u = {}
    for i in range(len(tasks)):
        for j in range(resources):
            u[i, j] = solver.IntVar(0, 1, f'u_{i}_{j}')

    # Variables z[i][t] for task i accessing some resource at time t
    z = {}
    for i in range(len(tasks)):
        for t in range(tasks[i][2]):
            z[i, t] = solver.IntVar(0, 1, f'z_{i}_{t}')

    # New variables y[i,j,t] to represent z[i,t] * u[i,j]
    y = {}
    for i in range(len(tasks)):
        for j in range(resources):
            for t in range(tasks[i][2]):
                y[i, j, t] = solver.IntVar(0, 1, f'y_{i}_{j}_{t}')

    # Constraints

    # Overlapping: check each pair of tasks to see if they are overlap time
    for i in range(len(tasks)):
        for ip in range(i + 1, len(tasks)):
            if check_overlap(tasks[i], tasks[ip]):
                for j in range(resources):
                    solver.Add(u[i, j] + u[ip, j] <= 1)

    # Symmetry breaking 1: Assign the tasks to resources if have r_max <= d_min (min of all tasks)
    d_min = min(task[2] for task in tasks)
    fixed_tasks = []
    for i in range(len(tasks)):
        if tasks[i][2] - tasks[i][1] <= d_min:
            fixed_tasks.append(i)
    # Assign each task in fixed_tasks to a resource
    for j, i in enumerate(fixed_tasks):
        if j < resources:
            solver.Add(u[i, j] == 1)

    # Symmetry breaking 2: if each task i has t in range(r_max, d_min), then z[i][t] = True
    for i in range(len(tasks)):
        for t in range(tasks[i][2] - tasks[i][1], tasks[i][0] + tasks[i][1]):
            if t < tasks[i][2]:  # Ensure we don't go beyond the task's deadline
                solver.Add(z[i, t] == 1)

    # D1 and D2: Task i should access exactly one resource
    for i in range(len(tasks)):
        solver.Add(sum(u[i, j] for j in range(resources)) == 1)

    # D3: A resource can only be held by one task at a time
    for j in range(resources):
        for t in range(max_time):
            solver.Add(sum(y[i, j, t] for i in range(len(tasks)) if t < tasks[i][2]) <= 1)

    # Linearization constraints for y[i,j,t] = z[i,t] * u[i,j]
    for i in range(len(tasks)):
        for j in range(resources):
            for t in range(tasks[i][2]):
                solver.Add(y[i, j, t] <= z[i, t])
                solver.Add(y[i, j, t] <= u[i, j])
                solver.Add(y[i, j, t] >= z[i, t] + u[i, j] - 1)

    # C3: Non-preemptive resource access
    for i in range(len(tasks)):
        solver.Add(sum(z[i, t] for t in range(tasks[i][0], tasks[i][2] - tasks[i][1] + 1)) == 1)

    # C4 and C5: Continuous execution
    for i in range(len(tasks)):
        for t in range(tasks[i][0], tasks[i][2] - tasks[i][1] + 1):
            solver.Add(sum(z[i, tp] for tp in range(t, min(t + tasks[i][1], tasks[i][2]))) >= 
                       tasks[i][1] * z[i, t])

    # Linking z and y variables
    for i in range(len(tasks)):
        for t in range(tasks[i][2]):
            solver.Add(z[i, t] == sum(y[i, j, t] for j in range(resources)))

    return solver, u, z, y

def validate_solution(tasks, solver, u, z, y, resources):
    task_resource = {}
    task_times = {}
    resource_usage = {j: [] for j in range(resources)}

    for i, task in enumerate(tasks):
        for j in range(resources):
            if u[i, j].solution_value() > 0.5:
                task_resource[i] = j
        
        task_times[i] = [t for t in range(task[0], task[2]) if z[i, t].solution_value() > 0.5]
        
        if task_resource.get(i) is not None:
            resource_usage[task_resource[i]].extend(task_times[i])

    # Check constraints
    for i, task in enumerate(tasks):
        if i not in task_resource:
            print_to_console_and_log(f"Error: Task {i} is not assigned to any resource")
            return False

        if task_times[i][0] < task[0]:
            print_to_console_and_log(f"Error: Task {i+1} starts before its release time")
            return False

        if task_times[i][-1] >= task[2]:
            print_to_console_and_log(f"Error: Task {i+1} finishes after its deadline")
            return False

        if len(task_times[i]) != task[1] or any(task_times[i][j+1] - task_times[i][j] != 1 for j in range(len(task_times[i])-1)):
            print_to_console_and_log(f"Error: Task {i+1} execution is not continuous or doesn't match execution time")
            return False

    for j, times in resource_usage.items():
        if len(times) != len(set(times)):
            print_to_console_and_log(f"Error: Resource {j+1} is used by multiple tasks at the same time")
            return False

    print_to_console_and_log("Solution is valid!")
    return True

def solve_with_timeout(tasks, resources, result_container, finished_event):
    try:
        # Move encoding and solving into this function
        solver, u, z, y = encode_problem_es3(tasks, resources)
        if not solver:
            result_container['status'] = 'ERROR'
            return

        solver.set_time_limit(time_budget * 1000)
        status = solver.Solve()

        if status == pywraplp.Solver.OPTIMAL or status == pywraplp.Solver.FEASIBLE:
            result_container['status'] = 'SAT'
            result_container['solver'] = solver
            result_container['u'] = u
            result_container['z'] = z
            result_container['y'] = y
        elif status == pywraplp.Solver.INFEASIBLE:
            result_container['status'] = 'UNSAT'
        else:
            result_container['status'] = 'TIMEOUT'
            
    except Exception as e:
        result_container['status'] = 'ERROR'
        result_container['error'] = str(e)
    
    finished_event.set()

def solve_es3(tasks, resources):
    result_container = {}
    finished_event = Event()
    
    start_time = time.time()
    solver_thread = Thread(target=solve_with_timeout, args=(tasks, resources, result_container, finished_event))
    solver_thread.start()
    
    # Wait for either completion or timeout
    finished = finished_event.wait(timeout=time_budget)
    solve_time = time.time() - start_time
    
    if not finished:
        solver_thread.join()  # Wait for thread to clean up
        return "TIMEOUT", solve_time, 0, 0
        
    if result_container.get('status') == 'SAT':
        solver = result_container['solver']
        u = result_container['u']
        z = result_container['z']
        y = result_container['y']
        
        print_to_console_and_log("Solution found.")
        for i in range(len(tasks)):
            for j in range(resources):
                if u[i, j].solution_value() > 0.5:
                    print_to_console_and_log(f"Task {i+1} is assigned to resource {j+1}")
            for t in range(tasks[i][0], tasks[i][2]):
                if z[i, t].solution_value() > 0.5:
                    print_to_console_and_log(f"Task {i+1} is accessing a resource at time {t}")
                    
        if not validate_solution(tasks, solver, u, z, y, resources):
            sys.exit(1)
            
        num_variables = solver.NumVariables()
        num_constraints = solver.NumConstraints()
        return "SAT", solve_time, num_variables, num_constraints
        
    elif result_container.get('status') == 'UNSAT':
        print_to_console_and_log("Problem is infeasible.")
        return "UNSAT", solve_time, 0, 0
    else:
        print_to_console_and_log("Solver failed or timed out.")
        return result_container.get('status', 'ERROR'), solve_time, 0, 0

def process_input_files(input_folder, resources=200):
    global id_counter, type

    # results = {}
    for filename in os.listdir(input_folder):
        if filename.endswith(".txt"):
            file_path = os.path.join(input_folder, filename)
            with open(file_path, 'r') as f:
                num_tasks = int(f.readline().strip())
                tasks = ast.literal_eval(f.readline().strip())
                print(f"tasks: {tasks}")

            print_to_console_and_log(f"Processing {filename}...")
            # res, solve_time, num_variables, num_clauses = solve_es3(tasks, num_tasks)
            res, solve_time, num_variables, num_clauses = solve_es3(tasks, resources)
            # results[filename] = {
            #     "result": res,
            #     "time": float(solve_time),
            #     "num_variables": num_variables,
            #     "num_clauses": num_clauses
            # }
            result_dict = {
                "ID": id_counter,
                "Problem": os.path.basename(filename),
                "Type": type,
                "Time": solve_time,
                "Result": res,
                "Variables": num_variables,
                "Clauses": num_clauses
            }
            write_to_xlsx(result_dict)
            id_counter += 1

    # return results

# Main execution
input_folder = "input/" + sys.argv[1]
# input_folder = "input/small"
process_input_files(input_folder)

log_file.close()
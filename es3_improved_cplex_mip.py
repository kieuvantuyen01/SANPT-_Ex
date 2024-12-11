import sys
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile
from threading import Thread, Event
import cplex
from itertools import product
import os
import ast
import time
from collections import defaultdict

time_budget = 600  # Set your desired time budget in seconds
type = "es3_improved_cplex_mip"
id_counter = 1

# Open the log file in append mode
log_file = open('console.log', 'a')

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'out/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except (BadZipFile, InvalidFileException, KeyError):
            print_to_console_and_log(f"Error: The existing file {excel_file_path} is not a valid Excel file. Creating a new one.")
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

    print_to_console_and_log(f"Result added to Excel file: {os.path.abspath(excel_file_path)}\n")


# Define a custom print function that writes to both console and log file
def print_to_console_and_log(*args, **kwargs):
    print(*args, **kwargs)
    print(*args, file = log_file, **kwargs)
    log_file.flush()

def check_overlap(task1, task2):
    # Suppose: task1 = (r1, e1, d1), task2 = (r2, e2, d2)
    # r1_min = r1, r1_max = d1 - e1, d1_min = r1 + e1, d1_max = d1
    # r2_min = r2, r2_max = d2 - e2, d2_min = r2 + e2, d2_max = d2
    # task1 and task2 are overlapped if: 
    # 1. d2_min >= r1_max and r2_max <= d1_min
    # 2. d1_min >= r2_max and r1_max <= d2_min
    # => r2 + e2 >= d1 - e1 and d2 - e2 <= r1 + e1 or r1 + e1 >= d2 - e2 and d1 - e1 <= r2 + e2
    if task2[0] + task2[1] > task1[2] - task1[1] and task2[2] - task2[1] < task1[0] + task1[1]:
        return True
    if task1[0] + task1[1] > task2[2] - task2[1] and task1[2] - task1[1] < task2[0] + task2[1]:
        return True
    return False

def encode_problem_es3(tasks, resources):
    cpx = cplex.Cplex()
    cpx.set_results_stream(None)
    cpx.set_log_stream(None)

    max_time = max(task[2] for task in tasks)

    # Variables u[i][j] for task i accessing resource j
    u = []
    for i in range(len(tasks)):
        for j in range(resources):
            u.append(f'u_{i}_{j}')

    # Variables z[i][t] for task i accessing some resource at time t
    z = []
    for i in range(len(tasks)):
        for t in range(tasks[i][2]):
            z.append(f'z_{i}_{t}')

    # Add variables
    cpx.variables.add(names=u + z, types=[cpx.variables.type.binary] * (len(u) + len(z)))

    # Use a dictionary to store unique constraints
    constraints = defaultdict(int)

    # Helper function to add constraints
    def add_constraint(ind, val, sense, rhs):
        # Sort the indices and values together
        sorted_pairs = tuple(sorted(zip(ind, val)))
        # Create a hashable representation of the constraint
        constraint = (sorted_pairs, sense, rhs)
        constraints[constraint] += 1
        if constraints[constraint] > 1:
            print(f"Duplicate constraint found: {constraint}")

    # Overlapping: check each pair of tasks to see if they are overlap time
    for i in range(len(tasks)):
        for ip in range(i + 1, len(tasks)):
            if check_overlap(tasks[i], tasks[ip]):
                for j in range(resources):
                    add_constraint([f'u_{i}_{j}', f'u_{ip}_{j}'], [1.0, 1.0], 'L', 1.0)

    # Symmetry breaking 1: Assign the tasks to resources if have r_max <= d_min (min of all tasks)
    d_min = min(task[2] for task in tasks)
    fixed_tasks = []
    for i in range(len(tasks)):
        if tasks[i][2] - tasks[i][1] <= d_min:
            fixed_tasks.append(i)
    # Assign each task in fixed_tasks to a resource
    for j, i in enumerate(fixed_tasks):
        if j < resources:
            add_constraint([f'u_{i}_{j}'], [1.0], 'E', 1.0)

    # Symmetry breaking 2: if each task i has t in range(r_max, d_min), then z[i][t] = True
    for i in range(len(tasks)):
        for t in range(tasks[i][2] - tasks[i][1], tasks[i][0] + tasks[i][1]):
            if t < tasks[i][2]:
                add_constraint([f'z_{i}_{t}'], [1.0], 'E', 1.0)

    # D1: Task i should not access two resources at the same time
    for i in range(len(tasks)):
        for j in range(resources):
            for jp in range(j + 1, resources):
                add_constraint([f'u_{i}_{j}', f'u_{i}_{jp}'], [1.0, 1.0], 'L', 1.0)

    # D2: Each task must get some resource
    for i in range(len(tasks)):
        add_constraint([f'u_{i}_{j}' for j in range(resources)], [1.0] * resources, 'E', 1.0)

    # D3: A resource can only be held by one task at a time
    for i in range(len(tasks)):
        for ip in range(i + 1, len(tasks)):
            for j in range(resources):
                for t in range(max(tasks[i][0], tasks[ip][0]), min(tasks[i][2], tasks[ip][2])):
                    add_constraint([f'z_{i}_{t}', f'u_{i}_{j}', f'z_{ip}_{t}', f'u_{ip}_{j}'], [1.0, 1.0, 1.0, 1.0], 'L', 3.0)

    # C3: Non-preemptive resource access
    for i in range(len(tasks)):
        ind = [f'z_{i}_{t}' for t in range(tasks[i][0], tasks[i][2] - tasks[i][1] + 1)]
        val = [1.0] * len(ind)
        add_constraint(ind, val, 'E', 1.0)

    # C4 and C5: Continuous execution
    for i in range(len(tasks)):
        # C41 and C42 remain the same
        for t in range(tasks[i][0] + 1, tasks[i][0] + tasks[i][1]):
            add_constraint([f'z_{i}_{tasks[i][0]}', f'z_{i}_{t}'], [-1.0, 1.0], 'G', 0.0)

        for t in range(tasks[i][0] + tasks[i][1], tasks[i][2]):
            add_constraint([f'z_{i}_{tasks[i][0]}', f'z_{i}_{t}'], [-1.0, -1.0], 'L', 0.0)

        # Modified C51 and C52
        for t in range(tasks[i][0], tasks[i][2] - tasks[i][1]):
            next_t = t + 1
            for tpp in range(next_t + 1, min(t + tasks[i][1] + 1, tasks[i][2])):
                add_constraint([f'z_{i}_{t}', f'z_{i}_{next_t}', f'z_{i}_{tpp}'], [1.0, -1.0, 1.0], 'G', 0.0)

            for tpp in range(t + tasks[i][1] + 1, tasks[i][2]):
                add_constraint([f'z_{i}_{t}', f'z_{i}_{next_t}', f'z_{i}_{tpp}'], [1.0, -1.0, -1.0], 'L', 1.0)

    # After generating all constraints, add the unique ones to the model
    for constraint, count in constraints.items():
        ind_val, sense, rhs = constraint
        ind, val = zip(*ind_val)
        try:
            cpx.linear_constraints.add(
                lin_expr=[cplex.SparsePair(ind=ind, val=val)],
                senses=[sense],
                rhs=[rhs]
            )
        except cplex.exceptions.CplexError as e:
            print(f"Error adding constraint: {constraint}")
            print(f"Error message: {str(e)}")

    print(f"Total number of unique constraints: {len(constraints)}")
    print(f"Constraints with duplicates: {sum(1 for count in constraints.values() if count > 1)}")

    return cpx, u, z

def validate_solution(tasks, cpx, u, z, resources):
    task_resource = {}
    task_times = {}
    resource_usage = {j: [] for j in range(resources)}

    for i, task in enumerate(tasks):
        for j in range(resources):
            if cpx.solution.get_values(f'u_{i}_{j}') > 0.5:
                task_resource[i] = j
        
        task_times[i] = [t for t in range(task[0], task[2]) if cpx.solution.get_values(f'z_{i}_{t}') > 0.5]
        
        if task_resource.get(i) is not None:
            resource_usage[task_resource[i]].extend(task_times[i])

    # Check constraints
    for i, task in enumerate(tasks):
        if i not in task_resource:
            print_to_console_and_log(f"Error: Task {i} is not assigned to any resource")
            return False

        if task_times[i][0] < task[0]:
            print_to_console_and_log(f"Error: Task {i+1} starts before its release time")
            return False

        if task_times[i][-1] >= task[2]:
            print_to_console_and_log(f"Error: Task {i+1} finishes after its deadline")
            return False

        if len(task_times[i]) != task[1] or any(task_times[i][j+1] - task_times[i][j] != 1 for j in range(len(task_times[i])-1)):
            print_to_console_and_log(f"Error: Task {i+1} execution is not continuous or doesn't match execution time")
            return False

    for j, times in resource_usage.items():
        if len(times) != len(set(times)):
            print_to_console_and_log(f"Error: Resource {j+1} is used by multiple tasks at the same time")
            return False

    print_to_console_and_log("Solution is valid!")
    return True

def solve_with_timeout(tasks, resources, result_container, finished_event):
    try:
        # Encode the problem
        cpx, u, z = encode_problem_es3(tasks, resources)
        if not cpx:
            result_container['status'] = 'ERROR'
            finished_event.set()
            return

        # Solve
        cpx.solve()
        
        # Store results
        result_container['cpx'] = cpx
        result_container['u'] = u
        result_container['z'] = z
        result_container['num_variables'] = cpx.variables.get_num()
        result_container['num_constraints'] = cpx.linear_constraints.get_num()
        
        status = cpx.solution.get_status()
        status_string = cpx.solution.get_status_string()
        print_to_console_and_log(f"Solution status: {status} ({status_string})")

        if status == cpx.solution.status.optimal:
            print_to_console_and_log("Optimal solution found.")
            result_container['status'] = "SAT"
        elif status == cpx.solution.status.feasible:
            print_to_console_and_log("Feasible solution found.")
            result_container['status'] = "SAT"
        elif status == cpx.solution.status.MIP_optimal:
            print_to_console_and_log("MIP solution is optimal.")
            result_container['status'] = "SAT"
        elif status == cpx.solution.status.infeasible:
            print_to_console_and_log("Problem is infeasible.")
            result_container['status'] = "UNSAT"
        else:
            print_to_console_and_log(f"Unexpected status: {status_string}")
            result_container['status'] = "UNKNOWN"
            
    except cplex.exceptions.CplexSolverError as e:
        print(f"Exception during solve: {e}")
        result_container['status'] = 'ERROR'
    except Exception as e:
        print(f"Unexpected error: {e}")
        result_container['status'] = 'ERROR'
    
    finished_event.set()

def solve_es3(tasks, resources):
    result_container = {}
    finished_event = Event()
    
    # Create thread and set daemon BEFORE starting
    solver_thread = Thread(target=solve_with_timeout, 
                         args=(tasks, resources, result_container, finished_event),
                         daemon=True) 
    start_time = time.time()
    solver_thread.start()
    
    # Wait for either completion or timeout
    finished = finished_event.wait(timeout=time_budget)
    solve_time = time.time() - start_time
    
    if not finished:
        print_to_console_and_log("Solver timed out.")
        if 'cpx' in result_container:
            try:
                result_container['cpx'].end()
            except:
                pass
                
        # Give thread 5 seconds max to cleanup
        solver_thread.join(timeout=5)
        if solver_thread.is_alive():
            print_to_console_and_log("Warning: Solver thread force terminated")
            # Thread will be terminated automatically since it's a daemon
        
        return "TIMEOUT", solve_time, 0, 0

    num_variables = result_container.get('num_variables', 0)
    num_constraints = result_container.get('num_constraints', 0)
    
    print_to_console_and_log(f"Solve time: {solve_time}")
    print_to_console_and_log(f"Num of variables: {num_variables}")
    print_to_console_and_log(f"Num of constraints: {num_constraints}")

    if result_container['status'] == "SAT":
        cpx = result_container['cpx']
        u = result_container['u']
        z = result_container['z']
        
        for i in range(len(tasks)):
            for j in range(resources):
                if cpx.solution.get_values(f'u_{i}_{j}') > 0.5:
                    print_to_console_and_log(f"Task {i+1} is assigned to resource {j+1}")
            for t in range(tasks[i][0], tasks[i][2]):
                if cpx.solution.get_values(f'z_{i}_{t}') > 0.5:
                    print_to_console_and_log(f"Task {i+1} is accessing a resource at time {t}")
        
        if not validate_solution(tasks, cpx, u, z, resources):
            sys.exit(1)

    return result_container['status'], solve_time, num_variables, num_constraints

def process_input_files(input_folder, resources=200):
    global id_counter, type

    # results = {}
    for filename in os.listdir(input_folder):
        if filename.endswith(".txt"):
            file_path = os.path.join(input_folder, filename)
            with open(file_path, 'r') as f:
                num_tasks = int(f.readline().strip())
                tasks = ast.literal_eval(f.readline().strip())
                print(f"tasks: {tasks}")

            print_to_console_and_log(f"Processing {filename}...")
            # res, solve_time, num_variables, num_clauses = solve_es3(tasks, num_tasks)
            res, solve_time, num_variables, num_clauses = solve_es3(tasks, resources)
            # results[filename] = {
            #     "result": res,
            #     "time": float(solve_time),
            #     "num_variables": num_variables,
            #     "num_clauses": num_clauses
            # }
            result_dict = {
                "ID": id_counter,
                "Problem": os.path.basename(filename),
                "Type": type,
                "Time": solve_time,
                "Result": res,
                "Variables": num_variables,
                "Clauses": num_clauses
            }
            write_to_xlsx(result_dict)
            id_counter += 1

    # return results

# Main execution
input_folder = "input/" + sys.argv[1]
# input_folder = "input/long_duration"
process_input_files(input_folder)

log_file.close()

